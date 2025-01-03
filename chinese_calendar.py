from datetime import datetime, timedelta
from lunar_python import Lunar
import calendar
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
from openpyxl.styles.numbers import FORMAT_TEXT
import requests
from PIL import Image, ImageDraw, ImageFont
from openpyxl.drawing.image import Image as XLImage
import io
import json
import os
import win32com.client
import time

class ChineseCalendar:
    def __init__(self, year, month, config_file='config.json'):
        self.year = year
        self.month = month
        
        # 加载配置文件
        self.config = self.load_config(config_file)
        
        # 从配置文件加载节假日
        self.holidays = self.config['custom_holidays']['solar']
        self.lunar_holidays = self.config['custom_holidays']['lunar']
        
        # 添加农历月份和日期的中文表示
        self.lunar_month_names = {
            1: "正月", 2: "二月", 3: "三月", 4: "四月", 5: "五月", 6: "六月",
            7: "七月", 8: "八月", 9: "九月", 10: "十月", 11: "冬月", 12: "腊月"
        }
        self.lunar_day_names = {
            1: "初一", 2: "初二", 3: "初三", 4: "初四", 5: "初五",
            6: "初六", 7: "初七", 8: "初八", 9: "初九", 10: "初十",
            11: "十一", 12: "十二", 13: "十三", 14: "十四", 15: "十五",
            16: "十六", 17: "十七", 18: "十八", 19: "十九", 20: "二十",
            21: "廿一", 22: "廿二", 23: "廿三", 24: "廿四", 25: "廿五",
            26: "廿六", 27: "廿七", 28: "廿八", 29: "廿九", 30: "三十"
        }
        
        # 初始化节假日数据缓存
        self._holiday_data_cache = None
        # 获取节假日数据
        self.holiday_data = self.get_holiday_data()
        # 创建"休"字图片
        self.rest_image = self.create_rest_mark()

    def load_config(self, config_file):
        """加载配置文件"""
        try:
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                # 使用默认配置
                return {
                    "column_width": 10.5,
                    "year": 2025,
                    "custom_holidays": {
                        "solar": {
                            "0101": "元旦",
                            "0501": "劳动节",
                            "1001": "国庆节"
                        },
                        "lunar": {
                            "0101": "春节",
                            "0115": "元宵节",
                            "0505": "端午节",
                            "0815": "中秋节"
                        }
                    }
                }
        except Exception as e:
            print(f"加载配置文件失败: {e}，将使用默认配置")
            return {
                "column_width": 10.5,
                "year": 2025,
                "custom_holidays": {
                    "solar": {
                        "0101": "元旦",
                        "0501": "劳动节",
                        "1001": "国庆节"
                    },
                    "lunar": {
                        "0101": "春节",
                        "0115": "元宵节",
                        "0505": "端午节",
                        "0815": "中秋节"
                    }
                }
            }

    def get_holiday_data(self):
        """获取节假日数据"""
        # 如果已经有缓存数据，直接返回
        if self._holiday_data_cache is not None:
            return self._holiday_data_cache
            
        url = "https://cdn.jsdelivr.net/npm/chinese-days/dist/chinese-days.json"
        try:
            response = requests.get(url)
            data = response.json()
            print("成功获取在线节假日数据")
            # 缓存数据
            self._holiday_data_cache = data
            return data
        except Exception as e:
            print(f"获取在线节假日数据失败: {e}")
            print("将仅使用配置文件中的节假日")
            default_data = {"holidays": {}, "workdays": {}}
            self._holiday_data_cache = default_data
            return default_data

    def is_holiday(self, date):
        """判断是否为节假日，并返回节假日名称"""
        # 1. 检查在线节假日数据（法定节假日）
        date_str = date.strftime("%Y-%m-%d")
        if date_str in self.holiday_data["holidays"]:
            holiday_info = self.holiday_data["holidays"][date_str]
            return True, holiday_info.split(",")[1]  # 返回中文名称
            
        # 2. 检查配置文件中的自定义公历节日
        date_str = date.strftime("%m%d")
        if date_str in self.holidays:
            return True, self.holidays[date_str]
            
        # 3. 检查配置文件中的自定义农历节日
        lunar_date = self.get_lunar_date(date)
        lunar_str = f"{lunar_date.lunar_month:02d}{lunar_date.lunar_day:02d}"
        if lunar_str in self.lunar_holidays:
            return True, self.lunar_holidays[lunar_str]
            
        return False, ""

    def get_lunar_date(self, solar_date):
        """将公历日期转换为农历日期"""
        lunar = Lunar.fromDate(solar_date)
        # 创建一个类来模拟 ZhDate 的接口
        class LunarDate:
            def __init__(self, lunar):
                self.lunar_month = lunar.getMonth()
                self.lunar_day = lunar.getDay()
                # 获取节气（如果当天是节气的话）
                jieqi = lunar.getJieQi()
                self.solar_term = jieqi if jieqi else None
        return LunarDate(lunar)

    def get_holiday(self, date):
        """获取节假日信息"""
        # 检查是否是节假日
        is_holiday_day, holiday_name = self.is_holiday(date)
        if is_holiday_day:
            return holiday_name
            
        return ""

    def get_lunar_date_str(self, lunar_date):
        """将农历日期转换为中文格式"""
        # 如果是节气，返回节气名称
        if lunar_date.solar_term:
            return lunar_date.solar_term
            
        # 确保月份为正数
        month = abs(lunar_date.lunar_month)
        if month == 0:  # 处理月份为0的特殊情况
            month = 1
        month_str = self.lunar_month_names[month]
        
        # 确保日期为正数
        day = abs(lunar_date.lunar_day)
        if day == 0:  # 处理日期为0的特殊情况
            day = 1
        
        # 如果是初一，只显示月份
        if day == 1:
            return month_str
        else:
            return self.lunar_day_names[day]

    def generate_month_calendar(self):
        """生成指定月份的日历"""
        print(f"\n{self.year}年{self.month}月\n")
        print("一  二  三  四  五  六  日")
        
        # 获取当月第一天
        first_day = datetime(self.year, self.month, 1)
        # 获取当月天数
        _, last_day = calendar.monthrange(self.year, self.month)
        
        # 打印日历前的空格
        week_day = first_day.weekday()
        print("   " * week_day, end="")
        
        # 打印日历
        current_day = first_day
        while current_day.month == self.month:
            lunar_date = self.get_lunar_date(current_day)
            holiday = self.get_holiday(current_day)
            
            date_str = f"{current_day.day:2d}"
            lunar_str = f"({lunar_date.lunar_day})"
            
            print(f"{date_str:<2}", end=" ")
            
            if (current_day.weekday() + 1) % 7 == 0:
                print()  # 换行
                
            current_day += timedelta(days=1)
        print("\n")

    def create_rest_mark(self):
        """创建'休'字图片"""
        # 从配置文件获取休字标记的设置
        rest_config = self.config.get('layout', {}).get('rest_mark', {})
        img_size = (
            rest_config.get('width', 20),
            rest_config.get('height', 20)
        )
        
        # 创建一个透明背景的图片，使用2倍大小以实现抗锯齿效果
        scale = 2
        img = Image.new('RGBA', 
                       (img_size[0] * scale, img_size[1] * scale), 
                       (255, 255, 255, 0))
        draw = ImageDraw.Draw(img)
        
        # 从配置文件获取字体设置
        font_name = rest_config.get('font_name', "华文细黑")
        font_size = rest_config.get('font_size', 14) * scale  # 字体也放大2倍
        
        try:
            # 尝试使用系统字体名称
            font = ImageFont.truetype(font_name, font_size)
        except:
            try:
                # 如果失败，尝试使用字体文件路径
                font = ImageFont.truetype(f"{font_name}.ttf", font_size)
            except:
                try:
                    # 再尝试使用STXIHEI.TTF
                    font = ImageFont.truetype("STXIHEI.TTF", font_size)
                except:
                    print("警告：无法加载指定字体，将使用默认字体")
                    # 如果还是失败，使用系统默认的中文字体
                    font = ImageFont.load_default()
        
        # 从配置文件获取颜色设置
        color_str = rest_config.get('color', "008000")
        # 将颜色代码转换为RGB值
        r = int(color_str[0:2], 16)
        g = int(color_str[2:4], 16)
        b = int(color_str[4:6], 16)
        
        # 获取文字偏移量
        text_offset_x = rest_config.get('text_offset_x', 3) * scale
        text_offset_y = rest_config.get('text_offset_y', 2) * scale
        
        # 绘制"休"字
        draw.text((text_offset_x, text_offset_y), "休", 
                 font=font, fill=(r, g, b, 255))
        
        # 将图片缩小回原始大小，这样可以获得抗锯齿效果
        img = img.resize(img_size, Image.LANCZOS)
        
        # 将图片保存到内存中
        img_byte_arr = io.BytesIO()
        img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)
        
        # 保存图片数据的副本
        self._rest_image_data = img_byte_arr.getvalue()
        
        return self._rest_image_data

    def get_rest_image(self):
        """获取休字图片的新实例"""
        if not hasattr(self, '_rest_image_data'):
            self.create_rest_mark()
        return io.BytesIO(self._rest_image_data)

    def offset_image(self, img, col, row):
        """精确设置图片位置"""
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU

        # 转换像素到EMU（English Metric Units）
        p2e = pixels_to_EMU
        h, w = img.height, img.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))

        # 从配置文件获取偏移设置
        rest_config = self.config.get('layout', {}).get('rest_mark', {})
        pixels_right = rest_config.get('offset_x', 44)  # 默认向右偏移44像素
        pixels_down = rest_config.get('offset_y', 0)    # 默认向下偏移0像素
        
        marker = AnchorMarker(
            col=col, 
            colOff=p2e(pixels_right),
            row=row, 
            rowOff=p2e(pixels_down)
        )
        img.anchor = OneCellAnchor(_from=marker, ext=size)

    def add_vba_macro(self, filename):
        """使用win32com添加VBA宏代码"""
        try:
            # 确保文件是.xlsx格式
            if filename.endswith('.xlsm'):
                temp_filename = filename[:-5] + '.xlsx'
            else:
                temp_filename = filename
                filename = filename[:-5] + '.xlsm'
            
            # 创建Excel应用程序实例
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            try:
                # 打开工作簿
                wb = excel.Workbooks.Open(os.path.abspath(temp_filename))
                
                try:
                    # 尝试访问VBA项目
                    vba_module = wb.VBProject.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
                except Exception as e:
                    if "不信任到 Visual Basic Project 的程序连接" in str(e):
                        print("\n需要修改Excel安全设置才能添加宏代码：")
                        print("1. 打开Excel")
                        print("2. 点击'文件' -> '选项' -> '信任中心'")
                        print("3. 点击'信任中心设置' -> 'VBA工程对象模型访问'")
                        print("4. 选中'信任对VBA工程对象模型的访问'")
                        print("5. 点击'确定'保存设置")
                        print("\n修改设置后，请重新运行程序。")
                        return False
                    else:
                        raise
                
                # 获取Excel文件所在的目录的绝对路径
                excel_dir = os.path.abspath(os.path.dirname(filename))
                svg_path = os.path.join(excel_dir, "休.svg")
                # 确保路径使用正斜杠
                svg_path = svg_path.replace('\\', '\\\\')
                
                vba_code = '''Sub 替换所有对象为图片()
    Dim ws As Worksheet
    Dim shp As Shape
    Dim picturePath As String
    
    ' 设置图片路径为绝对路径
    picturePath = "''' + svg_path + '''"
    
    ' 确认图片文件存在
    If Dir(picturePath) = "" Then
        MsgBox "找不到指定的图片文件！" & vbCrLf & picturePath, vbExclamation
        Exit Sub
    End If
    
    ' 遍历当前工作簿的所有工作表
    For Each ws In ThisWorkbook.Worksheets
        ' 如果工作表中有形状对象
        If ws.Shapes.Count > 0 Then
            ' 从后向前遍历所有形状（这样删除对象时不会影响索引）
            For i = ws.Shapes.Count To 1 Step -1
                ' 获取当前形状对象
                Set shp = ws.Shapes(i)
                
                ' 记录原始位置和大小
                Dim Left As Double: Left = shp.Left
                Dim Top As Double: Top = shp.Top
                Dim Width As Double: Width = shp.Width
                Dim Height As Double: Height = shp.Height
                
                ' 删除原始对象
                shp.Delete
                
                ' 插入新图片并设置位置和大小
                ws.Shapes.AddPicture _
                    Filename:=picturePath, _
                    LinkToFile:=False, _
                    SaveWithDocument:=True, _
                    Left:=Left, _
                    Top:=Top, _
                    Width:=Width, _
                    Height:=Height
            Next i
        End If
    Next ws
    
    MsgBox "所有对象已替换完成！", vbInformation
End Sub'''
                
                # 将代码写入模块
                vba_module.CodeModule.AddFromString(vba_code)
                
                # 另存为.xlsm文件
                wb.SaveAs(os.path.abspath(filename), FileFormat=52)  # 52 = xlOpenXMLWorkbookMacroEnabled
                wb.Close()
                
                print("已添加VBA宏代码")
                return True
            finally:
                # 确保Excel实例被关闭
                excel.Quit()
        except Exception as e:
            print(f"添加VBA宏代码失败: {e}")
            return False

    def close_excel_instances(self):
        """关闭所有Excel实例"""
        try:
            import win32com.client
            excel = win32com.client.GetObject(Class="Excel.Application")
            excel.Quit()
        except:
            pass  # 如果没有打开的Excel实例，会抛出异常，我们可以忽略

    def save_with_retry(self, wb, filename, max_retries=3, delay=1):
        """尝试保存文件，如果失败则重试"""
        import time
        
        for i in range(max_retries):
            try:
                # 尝试关闭Excel实例
                self.close_excel_instances()
                
                # 如果文件存在，先尝试删除
                if os.path.exists(filename):
                    try:
                        os.remove(filename)
                    except:
                        pass
                
                # 保存文件
                wb.save(filename)
                return True
            except PermissionError:
                if i < max_retries - 1:  # 如果不是最后一次尝试
                    print(f"保存失败，{delay}秒后重试...")
                    time.sleep(delay)
                else:
                    print(f"无法保存文件 {filename}，请确保文件未被其他程序打开")
                    return False
            except Exception as e:
                print(f"保存文件时出错: {e}")
                return False
        return False

    def generate_excel_calendar(self, filename="calendar.xlsx"):
        """生成Excel格式的日历"""
        # 先生成.xlsx文件
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.year}年{self.month}月"
        
        # 隐藏网格线
        ws.sheet_view.showGridLines = False

        # 插入一列在日历区域左边
        ws.insert_cols(1)

        # 从配置文件获取列宽并转换为字符数（1个字符约等于1.1个单位宽度）
        column_width = self.config.get('column_width', 10.5) * 1.1
        # 设置列宽
        for col in range(2, 9):  # B到H列
            ws.column_dimensions[chr(64 + col)].width = column_width
        # 设置A列宽度（左边空白列）
        ws.column_dimensions['A'].width = column_width

        # 从配置文件获取样式设置
        styles = self.config.get('styles', {})
        layout = self.config.get('layout', {})

        # 设置标题（从第2行开始）
        ws.merge_cells('B2:H2')
        ws['B2'] = f"{self.year}年{self.month}月"
        title_style = styles.get('title', {})
        ws['B2'].font = Font(
            name=title_style.get('font_name', '微软雅黑'),
            size=title_style.get('font_size', 16),
            bold=title_style.get('bold', True)
        )
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

        # 设置星期标题
        weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        weekday_style = styles.get('weekday', {})
        for col, day in enumerate(weekdays, 2):  # 从B列开始
            cell = ws.cell(row=3, column=col)
            cell.value = day
            cell.alignment = Alignment(horizontal='center', vertical='center')  # 水平垂直居中
            cell.font = Font(
                name=weekday_style.get('font_name', '微软雅黑'),
                size=weekday_style.get('font_size', 10),
                bold=weekday_style.get('bold', True)
            )
            cell.fill = PatternFill(
                start_color=weekday_style.get('fill_color', "CCCCCC"),
                end_color=weekday_style.get('fill_color', "CCCCCC"),
                fill_type="solid"
            )

        # 获取当月第一天
        first_day = datetime(self.year, self.month, 1)
        week_day = first_day.weekday()
        
        # 计算需要的总行数
        first_weekday = first_day.weekday()
        total_days = calendar.monthrange(self.year, self.month)[1]
        total_weeks = (first_weekday + total_days + 6) // 7
        total_rows = 3 + (total_weeks * 2)  # 标题行 + 星期行 + (每周2行)
        
        # 设置行高
        row_heights = layout.get('row_heights', {})
        ws.row_dimensions[2].height = row_heights.get('title', 30)  # 标题行
        ws.row_dimensions[3].height = row_heights.get('weekday', 20)  # 星期行
        
        # 设置日期和农历行高
        for r in range(4, total_rows + 1, 2):
            ws.row_dimensions[r].height = row_heights.get('date', 30)
            ws.row_dimensions[r+1].height = row_heights.get('lunar', 30)
        
        # 设置日历区域外边框（粗线）
        border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # 应用边框到整个日历区域
        for row in range(2, total_rows + 1):
            for col in range(2, 9):  # B到H列
                cell = ws.cell(row=row, column=col)
                if row == 2:  # 顶部边框
                    cell.border = Border(top=Side(style='thick'))
                elif row == total_rows:  # 底部边框
                    cell.border = Border(bottom=Side(style='thick'))
                if col == 2:  # 左侧边框
                    cell.border = Border(left=Side(style='thick'))
                elif col == 8:  # 右侧边框
                    cell.border = Border(right=Side(style='thick'))
                    
                # 设置角落的边框
                if row == 2 and col == 2:  # 左上角
                    cell.border = Border(left=Side(style='thick'), top=Side(style='thick'))
                elif row == 2 and col == 8:  # 右上角
                    cell.border = Border(right=Side(style='thick'), top=Side(style='thick'))
                elif row == total_rows and col == 2:  # 左下角
                    cell.border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
                elif row == total_rows and col == 8:  # 右下角
                    cell.border = Border(right=Side(style='thick'), bottom=Side(style='thick'))
        
        # 填充日历数据
        current_day = first_day
        row = 4  # 从第4行开始（紧接着星期标题）
        col = week_day + 3  # 从C列开始，所以要加3而不是2

        while current_day.month == self.month:
            # 日期单元格
            date_cell = ws.cell(row=row, column=col)
            # 农历单元格
            lunar_cell = ws.cell(row=row+1, column=col)
            
            # 获取农历和节日信息
            lunar_date = self.get_lunar_date(current_day)
            traditional_holiday = self.get_holiday(current_day)
            is_holiday_day, holiday_name = self.is_holiday(current_day)
            
            # 设置日期
            date_cell.value = current_day.day
            date_style = styles.get('date', {})
            date_cell.font = Font(
                name=date_style.get('font_name', 'DINPro-Bold'),
                size=date_style.get('font_size', 16)
            )
            date_cell.alignment = Alignment(horizontal='center', vertical='bottom')

            # 获取农历文本
            lunar_text = self.get_lunar_date_str(lunar_date)
            lunar_style = styles.get('lunar', {})

            if is_holiday_day:
                # 添加"休"字标记
                rest_config = self.config.get('layout', {}).get('rest_mark', {})
                use_shape = rest_config.get('use_shape', True)
                self.add_rest_mark(ws, col-1, row-1, use_shape)  # 因为Excel的行列索引从0开始
                
                # 添加节假日名称（绿色）
                holiday_text = holiday_name
                lunar_cell.value = f"{lunar_text}\n{holiday_text}"
                lunar_cell.font = Font(
                    name=lunar_style.get('font_name', '华文细黑'),
                    size=lunar_style.get('font_size', 8),
                    color=lunar_style.get('holiday_color', "008000")
                )
            else:
                lunar_cell.value = lunar_text
                # 如果是节气，使用橙色字体
                if lunar_date.solar_term:
                    lunar_cell.font = Font(
                        name=lunar_style.get('font_name', '华文细黑'),
                        size=lunar_style.get('font_size', 8),
                        color="FFA500"  # 橙色
                    )
                else:
                    lunar_cell.font = Font(
                        name=lunar_style.get('font_name', '华文细黑'),
                        size=lunar_style.get('font_size', 8)
                    )
            
            lunar_cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

            # 周末设置红色（如果不是节假日和节气）
            if col in [7, 8] and not is_holiday_day and not lunar_date.solar_term:  # 修改为7和8列
                date_cell.font = Font(
                    name=date_style.get('font_name', 'DINPro-Bold'),
                    size=date_style.get('font_size', 16),
                    color=date_style.get('weekend_color', "FF0000")
                )
                lunar_cell.font = Font(
                    name=lunar_style.get('font_name', '华文细黑'),
                    size=lunar_style.get('font_size', 8),
                    color=lunar_style.get('weekend_color', "FF0000")
                )

            # 移动到下一个单元格
            col += 1
            if col > 8:  # 修改为8
                col = 3  # 修改为3
                row += 2

            current_day += timedelta(days=1)

        # 保存为.xlsx文件
        if not self.save_with_retry(wb, filename):
            return False
        
        print(f"日历已保存到 {filename}")
        
        # 添加VBA宏代码并转换为.xlsm
        if filename.endswith('.xlsx'):
            xlsm_filename = filename[:-5] + '.xlsm'
            if self.add_vba_macro(filename):
                print(f"已生成启用宏的Excel文件: {xlsm_filename}")
        
        return True

    def generate_year_calendar(self, year=None, filename="calendar.xlsx"):
        """生成整年的日历，每个月一个工作表"""
        # 先生成.xlsx文件
        if year is None:
            year = self.config.get('year', 2025)
            
        wb = Workbook()
        
        # 删除默认创建的工作表
        wb.remove(wb.active)
        
        # 为每个月创建一个工作表
        for month in range(1, 13):
            # 创建新的日历实例，使用相同的配置文件路径
            cal = ChineseCalendar(year, month, config_file='config.json')
            
            # 创建工作表
            ws = wb.create_sheet(title=f"{month}月")
            
            # 隐藏网格线
            ws.sheet_view.showGridLines = False
            
            # 插入一列在日历区域左边
            ws.insert_cols(1)
            
            # 从配置文件获取列宽
            column_width = self.config.get('column_width', 10.5)
            # 设置列宽
            for col in range(2, 9):  # B到H列
                ws.column_dimensions[chr(64 + col)].width = column_width
            # 设置A列宽度（左边空白列）
            ws.column_dimensions['A'].width = column_width

            # 从配置文件获取样式和布局设置
            styles = self.config.get('styles', {})
            layout = self.config.get('layout', {})

            # 设置标题（从第2行开始）
            ws.merge_cells('B2:H2')
            ws['B2'] = f"{year}年{month}月"
            ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
            title_style = styles.get('title', {})
            ws['B2'].font = Font(
                name=title_style.get('font_name', '微软雅黑'),
                size=title_style.get('font_size', 16),
                bold=title_style.get('bold', True)
            )

            # 设置星期标题
            weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            weekday_style = styles.get('weekday', {})
            for col, day in enumerate(weekdays, 2):  # 从B列开始
                cell = ws.cell(row=3, column=col)
                cell.value = day
                cell.alignment = Alignment(horizontal='center', vertical='center')  # 水平垂直居中
                cell.font = Font(
                    name=weekday_style.get('font_name', '微软雅黑'),
                    size=weekday_style.get('font_size', 10),
                    bold=weekday_style.get('bold', True)
                )
                cell.fill = PatternFill(
                    start_color=weekday_style.get('fill_color', "CCCCCC"),
                    end_color=weekday_style.get('fill_color', "CCCCCC"),
                    fill_type="solid"
                )

            # 获取当月第一天
            first_day = datetime(year, month, 1)
            week_day = first_day.weekday()
            
            # 计算需要的总行数
            first_weekday = first_day.weekday()
            total_days = calendar.monthrange(year, month)[1]
            total_weeks = (first_weekday + total_days + 6) // 7
            total_rows = 3 + (total_weeks * 2)  # 标题行 + 星期行 + (每周2行)
            
            # 设置行高
            row_heights = layout.get('row_heights', {})
            ws.row_dimensions[2].height = row_heights.get('title', 30)  # 标题行
            ws.row_dimensions[3].height = row_heights.get('weekday', 20)  # 星期行
            
            # 设置日期和农历行高
            for r in range(4, total_rows + 1, 2):
                ws.row_dimensions[r].height = row_heights.get('date', 30)
                ws.row_dimensions[r+1].height = row_heights.get('lunar', 30)
            
            # 设置日历区域外边框（粗线）
            border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )
            
            # 应用边框到整个日历区域
            for row in range(2, total_rows + 1):
                for col in range(2, 9):  # B到H列
                    cell = ws.cell(row=row, column=col)
                    if row == 2:  # 顶部边框
                        cell.border = Border(top=Side(style='thick'))
                    elif row == total_rows:  # 底部边框
                        cell.border = Border(bottom=Side(style='thick'))
                    if col == 2:  # 左侧边框
                        cell.border = Border(left=Side(style='thick'))
                    elif col == 8:  # 右侧边框
                        cell.border = Border(right=Side(style='thick'))
                        
                    # 设置角落的边框
                    if row == 2 and col == 2:  # 左上角
                        cell.border = Border(left=Side(style='thick'), top=Side(style='thick'))
                    elif row == 2 and col == 8:  # 右上角
                        cell.border = Border(right=Side(style='thick'), top=Side(style='thick'))
                    elif row == total_rows and col == 2:  # 左下角
                        cell.border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
                    elif row == total_rows and col == 8:  # 右下角
                        cell.border = Border(right=Side(style='thick'), bottom=Side(style='thick'))
            
            # 填充日历数据
            current_day = first_day
            row = 4  # 从第4行开始（紧接着星期标题）
            col = week_day + 2   # 从C列开始，所以要加3而不是2

            while current_day.month == month:
                # 日期单元格
                date_cell = ws.cell(row=row, column=col)
                # 农历单元格
                lunar_cell = ws.cell(row=row+1, column=col)
                
                # 获取农历和节日信息
                lunar_date = cal.get_lunar_date(current_day)
                traditional_holiday = cal.get_holiday(current_day)
                is_holiday_day, holiday_name = cal.is_holiday(current_day)
                
                # 设置日期
                date_cell.value = current_day.day
                date_style = styles.get('date', {})
                date_cell.font = Font(
                    name=date_style.get('font_name', 'DINPro-Bold'),
                    size=date_style.get('font_size', 16)
                )
                date_cell.alignment = Alignment(horizontal='center', vertical='bottom')

                # 获取农历文本
                lunar_text = cal.get_lunar_date_str(lunar_date)
                lunar_style = styles.get('lunar', {})

                if is_holiday_day:
                    # 添加"休"字标记
                    rest_config = self.config.get('layout', {}).get('rest_mark', {})
                    use_shape = rest_config.get('use_shape', True)
                    self.add_rest_mark(ws, col-1, row-1, use_shape)  # 因为Excel的行列索引从0开始
                    
                    # 添加节假日名称（绿色）
                    holiday_text = holiday_name
                    lunar_cell.value = f"{lunar_text}\n{holiday_text}"
                    lunar_cell.font = Font(
                        name=lunar_style.get('font_name', '华文细黑'),
                        size=lunar_style.get('font_size', 8),
                        color=lunar_style.get('holiday_color', "008000")
                    )
                else:
                    lunar_cell.value = lunar_text
                    # 如果是节气，使用橙色字体
                    if lunar_date.solar_term:
                        lunar_cell.font = Font(
                            name=lunar_style.get('font_name', '华文细黑'),
                            size=lunar_style.get('font_size', 8),
                            color="FFA500"  # 橙色
                        )
                    else:
                        lunar_cell.font = Font(
                            name=lunar_style.get('font_name', '华文细黑'),
                            size=lunar_style.get('font_size', 8)
                        )
                
                lunar_cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

                # 周末设置红色（如果不是节假日和节气）
                if col in [7, 8] and not is_holiday_day and not lunar_date.solar_term:  # 修改为7和8列
                    date_cell.font = Font(
                        name=date_style.get('font_name', 'DINPro-Bold'),
                        size=date_style.get('font_size', 16),
                        color=date_style.get('weekend_color', "FF0000")
                    )
                    lunar_cell.font = Font(
                        name=lunar_style.get('font_name', '华文细黑'),
                        size=lunar_style.get('font_size', 8),
                        color=lunar_style.get('weekend_color', "FF0000")
                    )

                # 移动到下一个单元格
                col += 1
                if col > 8:  # 修改为8
                    col = 2  # 修改为2
                    row += 2

                current_day += timedelta(days=1)

        # 保存为.xlsx文件
        if not self.save_with_retry(wb, filename):
            return False
            
        print(f"全年日历已保存到 {filename}")
        
        # 添加VBA宏代码并转换为.xlsm
        if filename.endswith('.xlsx'):
            xlsm_filename = filename[:-5] + '.xlsm'
            if self.add_vba_macro(filename):
                print(f"已生成启用宏的Excel文件: {xlsm_filename}")
        
        return True

    def add_rest_mark_as_shape(self, ws, col, row):
        """使用文本框添加'休'字标记"""
        from openpyxl.drawing.shapes import Shape
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font, RegularTextRun
        from openpyxl.utils.units import pixels_to_EMU

        # 从配置文件获取休字标记的设置
        rest_config = self.config.get('layout', {}).get('rest_mark', {})
        
        # 创建一个文本框
        shape = Shape()
        shape.txBody = RegularTextRun("休")  # 设置文本内容
        
        # 设置文本框样式
        color_str = rest_config.get('color', "008000")
        font = Font(
            typeface=rest_config.get('font_name', "华文细黑"),
            sz=rest_config.get('font_size', 8) * 100,  # 字体大小需要乘以100
            color=color_str
        )
        
        # 设置文本属性
        rpr = CharacterProperties(latin=font, ea=font, cs=font)
        ppr = ParagraphProperties(algn='ctr')  # 居中对齐
        p = Paragraph(pPr=ppr, endParaRPr=rpr)
        shape.txBody.p_lst = [p]
        
        # 设置形状属性（无填充、无边框）
        shape.noFill = True  # 无填充
        shape.ln = None  # 无边框
        
        # 设置位置和大小
        width = rest_config.get('width', 15)
        height = rest_config.get('height', 15)
        pixels_right = rest_config.get('offset_x', 44)
        pixels_down = rest_config.get('offset_y', 0)
        
        # 转换为EMU单位
        p2e = pixels_to_EMU
        marker = AnchorMarker(
            col=col,
            colOff=p2e(pixels_right),
            row=row,
            rowOff=p2e(pixels_down)
        )
        size = XDRPositiveSize2D(p2e(width), p2e(height))
        
        # 设置锚点
        anchor = OneCellAnchor(_from=marker, ext=size)
        shape.anchor = anchor
        
        # 添加到工作表
        ws.add_shape(shape)

    def add_rest_mark(self, ws, col, row, use_shape=True):
        """添加'休'字标记，可选择使用文本框或图片"""
        if use_shape:
            self.add_rest_mark_as_shape(ws, col, row)
        else:
            # 使用图片方式
            img = XLImage(self.get_rest_image())
            img.width = 15
            img.height = 15
            self.offset_image(img, col, row)
            ws.add_image(img)

# 使用示例
if __name__ == "__main__":
    import argparse
    
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description='生成中国日历')
    parser.add_argument('--year', type=int, help='要生成的年份')
    parser.add_argument('--config', type=str, default='config.json', help='配置文件路径')
    args = parser.parse_args()
    
    try:
        # 创建日历实例
        cal = ChineseCalendar(args.year if args.year else 2025, 1, config_file=args.config)
        
        # 生成日历
        output_filename = f"calendar_{args.year if args.year else cal.config['year']}.xlsx"
        
        # 尝试生成日历
        if not cal.generate_year_calendar(args.year, output_filename):
            print("\n生成日历失败。")
            print("如果是因为Excel安全设置问题，请按照上述说明修改设置后重试。")
            print("如果问题仍然存在，请检查是否有其他Excel文件正在使用。")
            exit(1)
        else:
            print("\n日历生成成功！")
            print(f"1. Excel文件：{output_filename}")
            xlsm_filename = output_filename[:-5] + '.xlsm'
            if os.path.exists(xlsm_filename):
                print(f"2. 启用宏的Excel文件：{xlsm_filename}")
                print("\n要使用SVG图片替换功能：")
                print("1. 将'休.svg'文件放在Excel文件同目录下")
                print("2. 打开启用宏的Excel文件（.xlsm）")
                print("3. 点击'启用宏'")
                print("4. 运行'替换图片'宏")
    except Exception as e:
        print(f"\n发生错误: {e}")
        exit(1) 